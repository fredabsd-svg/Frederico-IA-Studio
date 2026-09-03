import os
import tempfile
import unittest
import zipfile

try:
    from openpyxl import load_workbook

    from kits import KitError
    from xlspro import CORES_GRAF, MOEDA_FMT, Planilha

    TEM_OPENPYXL = True
except Exception:  # openpyxl ausente no ambiente — o CI e o sandbox o instalam
    TEM_OPENPYXL = False


@unittest.skipUnless(TEM_OPENPYXL, "openpyxl não instalado")
class XlsProArtifactTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory(prefix="frederico-xlsx-")
        self.path = os.path.join(self.tmp.name, "fluxo_caixa.xlsx")

    def tearDown(self):
        self.tmp.cleanup()

    def _create_workbook(self):
        planilha = Planilha()
        ws = planilha.aba("Fluxo de Caixa")
        planilha.titulo(ws, "Fluxo de Caixa — Auditoria")
        rows = [
            ["Jan", 12500.0, 7100.0, "=B4-C4"],
            ["Fev", 13800.0, 7650.0, "=B5-C5"],
            ["Mar", 14250.0, 8000.0, "=B6-C6"],
            ["TOTAL", "=SUM(B4:B6)", "=SUM(C4:C6)", "=SUM(D4:D6)"],
        ]
        info = planilha.tabela(
            ws,
            ["Mês", "Entradas", "Saídas", "Saldo"],
            rows,
            inicio=3,
            moeda=["Entradas", "Saídas", "Saldo"],
            total=True,
        )
        planilha.grafico_barras(ws, info, "Mês", "Saldo", "Saldo mensal")
        planilha.salvar(self.path)

    def test_cria_xlsx_real_com_formula_estilo_congelamento_e_grafico(self):
        self._create_workbook()
        self.assertTrue(zipfile.is_zipfile(self.path))
        workbook = load_workbook(self.path, data_only=False)
        ws = workbook["Fluxo de Caixa"]
        self.assertEqual(ws["D4"].value, "=B4-C4")
        self.assertEqual(ws["D7"].value, "=SUM(D4:D6)")
        self.assertEqual(ws["B4"].number_format, MOEDA_FMT)
        self.assertEqual(ws.freeze_panes, "A4")
        self.assertFalse(ws.sheet_view.showGridLines)
        self.assertEqual(len(ws._charts), 1)

    def test_cinco_etapas_reabrem_e_preservam_o_mesmo_artefato(self):
        self._create_workbook()
        for stage in range(1, 6):
            workbook = load_workbook(self.path, data_only=False)
            audit = workbook.create_sheet(f"Revisão {stage}")
            audit["A1"] = "Etapa"
            audit["B1"] = stage
            audit["A2"] = "Artefato preservado"
            audit["B2"] = True
            workbook.save(self.path)

            reopened = load_workbook(self.path, data_only=False)
            self.assertEqual(reopened["Fluxo de Caixa"]["D4"].value, "=B4-C4")
            self.assertEqual(len(reopened["Fluxo de Caixa"]._charts), 1)
            for previous in range(1, stage + 1):
                self.assertIn(f"Revisão {previous}", reopened.sheetnames)

    def test_linha_maior_que_cabecalho_reprova_na_auditoria(self):
        """A v1 escrevia a célula extra FORA da tabela, sem formato e sem
        cabeçalho, e ninguém via. A v2 corta a linha na largura do cabeçalho
        (nada vaza) e a auditoria REPROVA, dizendo qual linha está errada."""
        planilha = Planilha()
        ws = planilha.aba("Dados")
        planilha.tabela(ws, ["A", "B"], [[1, 2, 3], [4, 5, 6]])
        with self.assertRaises(KitError) as erro:
            planilha.salvar(self.path)
        self.assertIn("linha-fora-do-cabecalho", str(erro.exception))
        workbook = load_workbook(self.path)
        self.assertIsNone(workbook["Dados"]["C2"].value)


    # ---------- v2: números, filtro, notas, impressão e conferência ----------
    def test_tabela_recebe_numeros_e_calcula_o_total(self):
        """String "R$ 412.300,00" vira TEXTO no Excel: nenhuma soma, gráfico ou
        tabela dinâmica funciona em cima dela. O kit recebe o número e aplica o
        formato pt-BR na célula."""
        p = Planilha(emissor="Escritório", titulo="DRE")
        ws = p.aba("DRE")
        p.tabela(ws, ["Trimestre", "Receita", "Margem"],
                 [["1T25", 412300, 0.277], ["2T25", 455900, 0.319]],
                 moeda=["Receita"], pct=["Margem"], total="soma")
        p.salvar(self.path)
        wb = load_workbook(self.path)
        aba = wb["DRE"]
        self.assertEqual(aba["B2"].value, 412300)          # NÚMERO, não texto
        self.assertEqual(aba["B2"].number_format, MOEDA_FMT)
        self.assertEqual(aba["C2"].number_format, "0.0%")
        # O TOTAL é calculado pelo kit; percentual não se soma.
        self.assertEqual(aba["A4"].value, "TOTAL")
        self.assertEqual(aba["B4"].value, 868200)
        self.assertIn(aba["C4"].value, (None, ""))

    def test_total_em_formula_deixa_a_planilha_viva(self):
        p = Planilha(emissor="Escritório")
        ws = p.aba("Dados")
        p.tabela(ws, ["Item", "Valor"], [["a", 10], ["b", 20]],
                 moeda=["Valor"], total="formula")
        p.salvar(self.path)
        aba = load_workbook(self.path)["Dados"]
        self.assertEqual(aba["B4"].value, "=SUM(B2:B3)")

    def test_filtro_no_cabecalho_e_cabecalho_repetido_na_impressao(self):
        p = Planilha()
        ws = p.aba("Dados")
        p.tabela(ws, ["Item", "Valor"], [["a", 10], ["b", 20]], moeda=["Valor"])
        p.salvar(self.path)
        aba = load_workbook(self.path)["Dados"]
        self.assertEqual(aba.auto_filter.ref, "A1:B3")
        self.assertEqual(aba.print_title_rows, "$1:$1")

    def test_impressao_ajusta_a_largura_da_folha(self):
        """12 colunas sem "ajustar à largura" imprimem metade numa folha e o
        resto noutra, com o cabeçalho só na primeira."""
        p = Planilha(emissor="Escritório")
        ws = p.aba("Larga")
        p.tabela(ws, ["C%d" % i for i in range(12)],
                 [[i for i in range(12)]], milhar=["C%d" % i for i in range(1, 12)])
        p.salvar(self.path)
        aba = load_workbook(self.path)["Larga"]
        self.assertEqual(aba.page_setup.fitToWidth, 1)
        self.assertEqual(aba.page_setup.fitToHeight, 0)
        # Sem `fitToPage` no sheetPr o Excel IGNORA o fitToWidth acima.
        self.assertTrue(aba.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(aba.page_setup.orientation, "landscape")

    def test_notas_ficam_na_ultima_aba(self):
        p = Planilha(emissor="Escritório")
        p.notas(["Fonte: balancetes", "Valores em R$"])
        p.aba("Depois")
        p.salvar(self.path)
        wb = load_workbook(self.path)
        self.assertEqual(wb.sheetnames[-1], "Notas")

    def test_painel_e_a_primeira_aba_e_a_area_de_impressao_cobre_os_graficos(self):
        """Os gráficos flutuam sobre a grade e não entram em `max_row`: uma área
        de impressão calculada só pelas células imprimia o painel SEM eles."""
        p = Planilha(emissor="Escritório", titulo="DRE")
        ws = p.aba("Dados")
        info = p.tabela(ws, ["Item", "Valor"], [["a", 10], ["b", 20]], moeda=["Valor"])
        p.painel(kpis=[(30, "Total", "moeda")],
                 graficos=[("barras", info, "Item", "Valor", "Valor por item")])
        p.salvar(self.path)
        wb = load_workbook(self.path)
        self.assertEqual(wb.sheetnames[0], "Resumo")
        import re as _re
        area = wb["Resumo"].print_area
        area = area[0] if isinstance(area, list) else area
        linha_final = int(_re.findall(r"\d+", str(area))[-1])
        self.assertGreaterEqual(linha_final, 20)

    def test_kpi_numerico_entra_como_numero_no_painel(self):
        p = Planilha(emissor="Escritório")
        p.aba("Dados")
        p.painel(kpis=[(1887900, "Receita", "moeda"), ("12 meses", "Vigência")])
        p.salvar(self.path)
        painel = load_workbook(self.path)["Resumo"]
        self.assertEqual(painel["B5"].value, 1887900)
        self.assertEqual(painel["B5"].number_format, MOEDA_FMT)

    def test_coluna_declarada_numerica_com_texto_reprova(self):
        """Um texto numa coluna de moeda faz a soma dar zero e o gráfico sair
        vazio, sem nenhum erro visível na tela."""
        p = Planilha()
        ws = p.aba("Dados")
        p.tabela(ws, ["Item", "Valor"], [["a", "R$ 10,00"]], moeda=["Valor"])
        with self.assertRaises(KitError) as erro:
            p.salvar(self.path)
        self.assertIn("coluna-numerica-com-texto", str(erro.exception))

    def test_placeholder_de_rascunho_reprova(self):
        p = Planilha()
        ws = p.aba("Dados")
        p.tabela(ws, ["Campo", "Valor"], [["Vencimento", "DD/MM/AAAA"]])
        with self.assertRaises(KitError) as erro:
            p.salvar(self.path)
        self.assertIn("placeholder", str(erro.exception))

    def test_coluna_inexistente_no_formato_falha_cedo(self):
        p = Planilha()
        ws = p.aba("Dados")
        with self.assertRaises(KitError):
            p.tabela(ws, ["Serviço", "Valor mensal"], [["a", 1]], moeda=["Valor"])

    def test_grafico_rejeita_contrato_invalido_com_erro_claro(self):
        planilha = Planilha()
        ws = planilha.aba("Dados")
        with self.assertRaisesRegex(ValueError, "dict retornado"):
            planilha.grafico_barras(ws, [["A", "B"]], "A", "B")

    def test_coluna_de_moeda_larga_o_suficiente_para_o_valor_formatado(self):
        # "R$ 15.015,00" tem 12 caracteres; a largura da coluna precisa comportar
        # o TEXTO EXIBIDO (não o valor bruto 15015.0), senão o Excel mostra ######.
        planilha = Planilha()
        ws = planilha.aba("V")
        planilha.tabela(ws, ["Item", "Total"],
                        [["Serviço A", 15015.0], ["Serviço B", 2670.0]],
                        moeda=["Total"])
        planilha.salvar(self.path)
        wb = load_workbook(self.path)
        self.assertGreaterEqual(wb["V"].column_dimensions["B"].width, 12)

    def test_segunda_tabela_respira_e_nao_rouba_o_congelamento(self):
        planilha = Planilha()
        ws = planilha.aba("V")
        planilha.titulo(ws, "Título")
        i1 = planilha.tabela(ws, ["A", "B"], [["x", 1.0]])
        i2 = planilha.tabela(ws, ["A", "B"], [["y", 2.0]])
        planilha.salvar(self.path)
        wb = load_workbook(self.path)
        ws2 = wb["V"]
        # título na linha 1 → 1ª tabela respira até a linha 3 (linha 2 em branco)
        self.assertEqual(i1["r0"], 3)
        # 2ª tabela deixa 1 linha em branco depois da 1ª (não cola no conteúdo)
        self.assertGreater(i2["r0"], i1["r1"] + 1)
        # congelamento permanece no cabeçalho da 1ª tabela (não pula p/ a 2ª)
        self.assertEqual(ws2.freeze_panes, "A4")

    def test_largura_de_coluna_compartilhada_pega_a_maior(self):
        # Duas tabelas empilhadas nas mesmas colunas: a coluna tem UMA largura;
        # ela deve caber o maior conteúdo das duas, não encolher para a última.
        planilha = Planilha()
        ws = planilha.aba("V")
        planilha.tabela(ws, ["A", "B"], [["Mesa de escritório grande", 1.0]])
        planilha.tabela(ws, ["A", "B"], [["curto", 2.0]])
        planilha.salvar(self.path)
        wb = load_workbook(self.path)
        self.assertGreaterEqual(wb["V"].column_dimensions["A"].width, len("Mesa de escritório grande"))


    def test_caractere_de_controle_nao_derruba_a_planilha(self):
        """O openpyxl levanta IllegalCharacterError ao gravar caractere de
        controle. Uma célula suja vinda de PDF ou CSV derrubava a geração
        inteira — agora ela é limpa antes de ser escrita."""
        p = Planilha()
        ws = p.aba("Da\x07dos")
        p.titulo(ws, "T\x00ítulo")
        p.tabela(ws, ["Produto", "Qtd"], [["A\x1fB", 3], ["C", 4]])
        p.salvar(self.path)
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        aba = wb[wb.sheetnames[0]]
        self.assertNotIn("\x07", wb.sheetnames[0])
        textos = [c.value for row in aba.iter_rows() for c in row if isinstance(c.value, str)]
        self.assertTrue(any("AB" == t for t in textos), textos)

    def test_nome_de_aba_invalido_e_corrigido(self):
        p = Planilha()
        ws = p.aba("Vendas/2025: resumo [final]")
        self.assertNotIn("/", ws.title)
        self.assertNotIn(":", ws.title)
        self.assertLessEqual(len(ws.title), 31)

    # ---------- aba-painel e identidade "Tinta & Latão" ----------
    def test_painel_vira_a_primeira_aba_com_os_kpis_e_o_carimbo(self):
        """Quem abre a planilha cai na PRIMEIRA aba: sem o painel, isso é a
        base de dados crua."""
        p = Planilha(emissor="Frederico Assessoria Contábil")
        ws = p.aba("Vendas")
        p.tabela(ws, ["Produto", "Total"], [["Malha", 5400.0], ["Oxford", 6000.0]],
                 moeda=["Total"])
        p.painel("Painel — Vendas 2025",
                 kpis=[("R$ 11,4 mil", "Faturamento"), (2, "Produtos")],
                 atualizado="27/07/2026")
        p.salvar(self.path)
        wb = load_workbook(self.path)
        self.assertEqual(wb.sheetnames[0], "Resumo")
        textos = [c.value for row in wb["Resumo"].iter_rows()
                  for c in row if c.value is not None]
        self.assertIn("Painel — Vendas 2025", textos)
        self.assertIn("R$ 11,4 mil", textos)
        self.assertIn("FATURAMENTO", textos)
        carimbo = [t for t in textos if isinstance(t, str) and "ATUALIZADO EM" in t]
        self.assertEqual(len(carimbo), 1)
        self.assertIn("FREDERICO ASSESSORIA CONTÁBIL", carimbo[0])

    def test_grafico_no_painel_le_os_dados_da_aba_de_origem(self):
        """O gráfico mora no painel, mas as células referenciadas têm de ser as
        da aba de dados — senão sai vazio."""
        p = Planilha()
        ws = p.aba("Vendas")
        info = p.tabela(ws, ["Produto", "Total"],
                        [["Malha", 5400.0], ["Oxford", 6000.0], ["TOTAL", 11400.0]],
                        moeda=["Total"], total=True)
        painel = p.painel("Painel", kpis=[("R$ 11,4 mil", "Faturamento")])
        p.grafico_barras(painel, info, "Produto", "Total", "Total por produto",
                         anchor="B10")
        p.salvar(self.path)
        wb = load_workbook(self.path)
        self.assertEqual(len(wb["Resumo"]._charts), 1)
        self.assertEqual(len(wb["Vendas"]._charts), 0)
        ref = str(wb["Resumo"]._charts[0].series[0].val.numRef.f)
        self.assertIn("Vendas", ref)
        # a linha de TOTAL fica FORA do gráfico (senão achata as outras barras)
        self.assertTrue(ref.endswith("$3"), ref)

    def test_graficos_saem_com_as_cores_do_tema(self):
        """Regressão silenciosa: `DataPoint(graphicalProperties=...)` levanta
        TypeError (o argumento é `spPr`) e, com o erro engolido, a pizza saía
        com a paleta padrão do Excel."""
        p = Planilha()
        ws = p.aba("Vendas")
        info = p.tabela(ws, ["Produto", "Total"],
                        [["Malha", 5400.0], ["Oxford", 6000.0], ["Linho", 4700.0]],
                        moeda=["Total"])
        p.grafico_pizza(ws, info, "Produto", "Total", "Participação", anchor="E2")
        p.grafico_barras(ws, info, "Produto", "Total", "Total", anchor="E20")
        p.salvar(self.path)
        wb = load_workbook(self.path)
        pizza, barras = wb["Vendas"]._charts[0], wb["Vendas"]._charts[1]

        def cor(spPr):  # openpyxl devolve str ou ColorChoice conforme o caminho
            fill = spPr.solidFill
            if not isinstance(fill, str):
                fill = fill.srgbClr
            return fill if isinstance(fill, str) else fill.value

        pontos = pizza.series[0].data_points
        self.assertEqual(len(pontos), 3)
        self.assertEqual([cor(d.graphicalProperties) for d in pontos], CORES_GRAF[:3])
        self.assertEqual(cor(barras.series[0].graphicalProperties), CORES_GRAF[0])


if __name__ == "__main__":
    unittest.main()
