"""xlspro — kit de DESIGN para planilhas Excel (.xlsx) com openpyxl.

Instalado no sandbox do Frederico AI Studio, na mesma identidade
**"Tinta & Latão"** do docpro/pdfpro (a paleta, a escala e a formatação pt-BR
vêm de `kits.py`): cabeçalho em verde-tinta com filete de latão, zebra, bordas
horizontais discretas, formatos de número de verdade (R$, %, milhar, data),
linha de TOTAL, filtro no cabeçalho, congelamento, largura de coluna pelo texto
EXIBIDO, gráficos com as cores do tema, a **aba-painel** (KPIs + gráficos), a
aba de **Notas** e a **configuração de impressão** — sem ela o painel vaza para
uma segunda folha e a planilha impressa fica inutilizável.

Uso mínimo:
    from xlspro import Planilha
    p = Planilha(emissor="Meu Escritório", cliente="ACME LTDA", titulo="DRE 2025")
    ws = p.aba("DRE 2025")
    info = p.tabela(ws, ["Trimestre", "Receita", "Custos"], linhas,
                    moeda=["Receita", "Custos"], total="soma", filtro=True)
    p.painel(kpis=[(1887900, "Receita líquida", "moeda")],
             graficos=[("barras", info, "Trimestre", "Receita", "Receita por trimestre")])
    p.notas(["Fonte: balancetes conciliados", "Valores em R$"])
    rel = p.salvar("/workspace/outputs/dre-2025.xlsx")
    print("CONFERÊNCIA:", rel)      # {"ok": True, "abas": 3, "achados": []}

**Passe NÚMEROS, não strings formatadas.** Quem formata é a coluna
(`moeda=`, `pct=`, `milhar=`, `data=`); string com "R$" vira texto no Excel e
nenhuma soma, gráfico ou tabela dinâmica funciona em cima dela.
"""
import os
import re
import warnings
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.worksheet.properties import PageSetupProperties

import kits
from kits import (ESCALA, KitError, KitError as _KitError, achado,
                  achados_de_placeholder, e_numero, falha_se_grave, fmt,
                  limpa_texto, linha_de_total, normaliza_linhas, paleta_para,
                  relatorio, tipos_de_coluna)

#: Tipografia com fidelidade no cliente — a mesma do docpro.
F_SERIF = kits.TIPOGRAFIA["office"]["serif"]
F_SANS = kits.TIPOGRAFIA["office"]["sans"]
PALETA = kits.PALETA
#: Cores dos gráficos, na ordem das séries/fatias (openpyxl usa "RRGGBB").
CORES_GRAF = [c.lstrip("#") for c in kits.CORES_GRAF]

# Formatos de número. O código de moeda 416 é pt-BR: com ele o Excel mostra
# "R$ 1.234,56" em qualquer máquina, inclusive numa instalação em inglês. O
# negativo sai em VERMELHO e entre parênteses — a convenção do demonstrativo
# contábil, e o que evita um prejuízo passar despercebido numa coluna.
MOEDA_FMT = '[$R$-416] #,##0.00;[Red]([$R$-416] #,##0.00)'
PCT_FMT = '0.0%'
MILHAR_FMT = '#,##0;[Red](#,##0)'
DATA_FMT = 'dd/mm/aaaa'
FORMATOS = {"moeda": MOEDA_FMT, "pct": PCT_FMT, "milhar": MILHAR_FMT, "data": DATA_FMT}

#: Acima deste número de colunas a impressão vai para paisagem.
COLUNAS_PARA_PAISAGEM = 6


def _num(v):
    return e_numero(v)


def _texto_exibido(val, kind):
    """Comprimento aproximado do TEXTO que o Excel realmente mostra na célula,
    já considerando o formato de número. Sem isso, uma coluna de moeda cujo
    valor bruto é 15015.0 (5 chars) ganhava largura de 5 e aparecia como
    ###### — porque o exibido é "R$ 15.015,00" (12 chars)."""
    if isinstance(val, (date, datetime)):
        return 10
    if isinstance(val, bool) or not isinstance(val, (int, float)):
        return len(str(val))
    neg = 1 if val < 0 else 0
    if kind == "moeda":
        return len("R$ " + format(abs(val), ",.2f")) + neg
    if kind == "pct":
        return len(format(val * 100, ".1f")) + 1  # "15.6" + "%"
    if kind == "milhar":
        return len(format(abs(val), ",.0f")) + neg
    return len(str(val))


def _nome_de_aba(nome, padrao="Planilha"):
    # O Excel recusa nome de aba com : \ / ? * [ ] e limita a 31 caracteres.
    limpo = re.sub(r"[:\\/?*\[\]]", "-", limpa_texto(str(nome))).strip() or padrao
    return limpo[:31]


class Planilha:
    """Pasta de trabalho do kit. Sem abas no início — use `aba()`/`painel()`."""

    def __init__(self, cor_marca=None, emissor="", cliente="", titulo="",
                 tipografia="office"):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.pal = paleta_para(cor_marca)
        self.emissor = emissor
        self.cliente = cliente
        self.titulo_doc = titulo
        familia = kits.TIPOGRAFIA.get(tipografia) or kits.TIPOGRAFIA["office"]
        self.fonte = familia["sans"]
        self.fonte_titulo = familia["serif"]
        self._tabelas = []        # infos, para a auditoria e para o painel
        self._painel = None
        self._notas = None
        self._achados_geracao = []

    # ---------- abas ----------
    def aba(self, nome):
        ws = self.wb.create_sheet(title=_nome_de_aba(nome))
        ws.sheet_view.showGridLines = False
        return ws

    def titulo(self, ws, texto, colspan=6, linha=1):
        c = ws.cell(row=linha, column=1, value=limpa_texto(texto))
        c.font = Font(name=self.fonte_titulo, size=ESCALA["h1"], bold=True,
                      color=self.pal["tinta"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha,
                       end_column=max(1, colspan))
        ws.row_dimensions[linha].height = 26
        return linha + 1

    # ---------- tabela ----------
    def tabela(self, ws, cabecalho, linhas, inicio=None, moeda=None, pct=None,
               milhar=None, data=None, total=False, congelar=True, filtro=True,
               titulo=None, largura_max=60):
        """Escreve uma tabela estilizada com NÚMEROS de verdade.

        `moeda`/`pct`/`milhar`/`data` = listas de NOMES de coluna do cabeçalho;
        o kit aplica o formato pt-BR e alinha à direita. `total`:

          * `"soma"` ou `True` — o kit CALCULA a linha de total em Python;
          * `"formula"` — escreve `=SUM(...)`, para quem quer a planilha viva;
          * `False` — sem linha de total.

        Se a última linha que você passou já for um TOTAL, o kit reconhece e não
        acrescenta outra. `filtro=True` põe o auto-filtro no cabeçalho.

        Retorna dict com {ws, r0 (linha do cabeçalho), r1 (última linha de
        dados), c0, c1, cols, total} para uso em gráficos."""
        cabecalho = list(cabecalho)
        tipos = tipos_de_coluna(cabecalho, moeda, pct, milhar, data)
        linhas, achados = normaliza_linhas(cabecalho, linhas)
        self._achados_geracao.extend(achados)
        modo = str(total).lower() if total else ""
        ja_tem_total = bool(linhas) and str(linhas[-1][0]).strip().upper().startswith("TOTAL")
        formula = modo == "formula"
        if modo in ("true", "soma", "formula") and not ja_tem_total:
            linhas = linhas + [linha_de_total(cabecalho, linhas, tipos)]
        com_total = bool(modo) or ja_tem_total

        # Onde a tabela começa. Se `inicio` não vier: aba vazia → linha 1; caso
        # contrário deixa UMA linha de respiro em branco.
        if inicio:
            r0 = inicio
        elif ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            r0 = 1
        else:
            r0 = ws.max_row + 2
        if titulo:
            r0 = self.titulo(ws, titulo, colspan=len(cabecalho), linha=r0) + 1
        ncols = len(cabecalho)
        thin = Side(style="thin", color=self.pal["borda"])
        head_fill = PatternFill("solid", fgColor=self.pal["tinta"])
        for j, nome in enumerate(cabecalho, start=1):
            c = ws.cell(row=r0, column=j, value=limpa_texto(nome))
            c.font = Font(name=self.fonte, size=ESCALA["corpo"], bold=True,
                          color=self.pal["branco"])
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = Border(bottom=Side(style="medium", color=self.pal["latao"]))
        ws.row_dimensions[r0].height = 20

        n = len(linhas)
        zebra_fill = PatternFill("solid", fgColor=self.pal["suave"])
        for i, linha in enumerate(linhas):
            r = r0 + 1 + i
            eh_total = com_total and i == n - 1
            zebra = (not eh_total) and n >= 6 and (i % 2 == 1)
            for j, val in enumerate(linha, start=1):
                tipo = tipos.get(j - 1)
                if eh_total and formula and tipo in ("moeda", "milhar") and j > 1:
                    letra = get_column_letter(j)
                    val = "=SUM(%s%d:%s%d)" % (letra, r0 + 1, letra, r - 1)
                else:
                    val = limpa_texto(val) if isinstance(val, str) else val
                c = ws.cell(row=r, column=j, value=val)
                c.font = Font(name=self.fonte, size=ESCALA["corpo"], bold=eh_total,
                              color=self.pal["tinta"] if eh_total else self.pal["corpo"])
                if eh_total:
                    c.fill = zebra_fill
                    c.border = Border(top=Side(style="medium", color=self.pal["latao"]),
                                      bottom=thin)
                else:
                    if zebra:
                        c.fill = zebra_fill
                    c.border = Border(bottom=thin)
                if tipo in FORMATOS:
                    c.number_format = FORMATOS[tipo]
                numerica = _num(val) or (isinstance(val, str) and val.startswith("="))
                c.alignment = Alignment(horizontal="right" if numerica else "left",
                                        vertical="center")

        # Largura pelo conteúdo EXIBIDO, com teto — coluna de 200 caracteres
        # empurra a tabela para fora da folha impressa.
        def _kind(j):
            return tipos.get(j)
        for j, nome in enumerate(cabecalho, start=1):
            kind = _kind(j - 1)
            largura = len(str(nome)) + 2  # cabeçalho em negrito ocupa mais
            for linha in linhas:
                if j - 1 < len(linha):
                    largura = max(largura, _texto_exibido(linha[j - 1], kind))
            L = get_column_letter(j)
            nova = min(largura_max, max(10, largura + 3))
            # No Excel a largura é UMA por coluna: se outra tabela já usa esta
            # coluna acima, mantém a maior das duas.
            atual = ws.column_dimensions[L].width
            ws.column_dimensions[L].width = max(atual or 0, nova)

        # Congela o cabeçalho só da PRIMEIRA tabela da aba (freeze_panes é
        # propriedade única da aba) e repete-o na impressão.
        if congelar and ws.freeze_panes is None:
            ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
            ws.print_title_rows = "%d:%d" % (r0, r0)
        if filtro and ws.auto_filter.ref is None:
            ultima = r0 + n - (1 if com_total else 0)
            ws.auto_filter.ref = "A%d:%s%d" % (r0, get_column_letter(ncols),
                                               max(r0, ultima))
        info = {"ws": ws, "r0": r0, "r1": r0 + n, "c0": 1, "c1": ncols,
                "cols": list(cabecalho), "total": com_total, "tipos": tipos}
        self._tabelas.append(info)
        return info

    def _idx(self, info, nome):
        # Mensagem clara quando o 2º argumento do gráfico não é o dict retornado
        # por p.tabela(...) — evita um TypeError opaco e ajuda a corrigir.
        if not isinstance(info, dict) or "cols" not in info:
            raise KitError(
                "grafico_*: passe como 2º argumento o dict retornado por "
                "p.tabela(...) (o 'info'), não uma lista/tabela crua.")
        cols = info["cols"]
        if nome not in cols:
            raise KitError(
                "grafico_*: coluna '%s' não existe no cabecalho %s" % (nome, cols))
        return cols.index(nome) + 1

    # ---------- aba-painel ----------
    def painel(self, titulo=None, kpis=None, graficos=None, nome="Resumo",
               atualizado=None, largura_cols=12):
        """Cria a aba-resumo na PRIMEIRA posição: título, carimbo do emissor,
        cartões de KPI (3 por linha) e os gráficos ancorados numa grade abaixo.

        Existe porque quem abre a planilha cai na primeira aba: sem o painel,
        isso é a base de dados crua. `kpis` = `[(valor, rótulo)]` ou
        `[(valor, rótulo, tipo)]` com tipo `moeda | pct | milhar | texto` — o
        kit formata o número. `graficos` = `[(tipo, info, categoria, valor,
        titulo)]`, com tipo `barras | linhas | pizza`."""
        titulo = titulo or self.titulo_doc or "Painel"
        ws = self.wb.create_sheet(title=_nome_de_aba(nome, "Resumo"), index=0)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 2.5
        ncols = max(8, largura_cols)
        for j in range(2, ncols + 2):
            ws.column_dimensions[get_column_letter(j)].width = 12
        c = ws.cell(row=2, column=2, value=limpa_texto(titulo))
        c.font = Font(name=self.fonte_titulo, size=17, bold=True, color=self.pal["tinta"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=ncols + 1)
        ws.row_dimensions[2].height = 30
        if atualizado is None:
            atualizado = fmt.data()
        carimbo = " · ".join(x for x in (
            (self.emissor or "").upper(),
            (self.cliente or "").upper(),
            ("ATUALIZADO EM " + str(atualizado)) if atualizado else "") if x)
        if carimbo:
            c3 = ws.cell(row=3, column=2, value=carimbo)
            c3.font = Font(name=self.fonte, size=8, bold=True, color=self.pal["latao"])
            ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=ncols + 1)
        filete = Side(style="medium", color=self.pal["tinta"])
        for j in range(2, ncols + 2):
            ws.cell(row=3, column=j).border = Border(bottom=filete)
        linha = self._cartoes(ws, kpis, 5) if kpis else 5
        for i, spec in enumerate(list(graficos or [])):
            tipo, info, categoria, valor = list(spec)[:4]
            titulo_g = spec[4] if len(spec) > 4 else ""
            # Grade de 2 colunas de gráfico, cada um com ~8 linhas de altura.
            coluna = "B" if i % 2 == 0 else "J"
            ancora = "%s%d" % (coluna, linha + (i // 2) * 17)
            self._grafico_por_tipo(tipo, ws, info, categoria, valor, titulo_g, ancora)
        self._painel = ws
        return ws

    def _cartoes(self, ws, kpis, r0):
        """Cartões de KPI, 3 por linha. Seis cartões numa fileira só ficam com
        uma coluna e meia cada e nenhum valor cabe."""
        fill = PatternFill("solid", fgColor=self.pal["suave"])
        topo = Side(style="thick", color=self.pal["latao"])
        col = 2
        for i, item in enumerate(list(kpis)):
            item = list(item)
            valor, rotulo = item[0], item[1]
            tipo = item[2] if len(item) > 2 else None
            tipo = None if tipo in (None, "texto", "num") else tipo
            if i and i % 3 == 0:
                r0 += 4
                col = 2
            c0, c1 = col, col + 2
            for rr in range(r0, r0 + 3):
                for cc in range(c0, c1 + 1):
                    cel = ws.cell(row=rr, column=cc)
                    cel.fill = fill
                    if rr == r0:
                        cel.border = Border(top=topo)
            ws.merge_cells(start_row=r0, start_column=c0, end_row=r0 + 1, end_column=c1)
            cv = ws.cell(row=r0, column=c0)
            # Número entra como NÚMERO e ganha o formato da célula: assim o
            # painel continua somável e o valor não vira texto decorativo.
            if e_numero(valor) and tipo in FORMATOS:
                cv.value = valor
                cv.number_format = FORMATOS[tipo]
            else:
                cv.value = limpa_texto(kits.formata_valor(valor, tipo))
            cv.font = Font(name=self.fonte_titulo, size=16, bold=True,
                           color=self.pal["tinta"])
            cv.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(start_row=r0 + 2, start_column=c0, end_row=r0 + 2, end_column=c1)
            cr = ws.cell(row=r0 + 2, column=c0, value=limpa_texto(str(rotulo)).upper())
            cr.font = Font(name=self.fonte, size=7.5, bold=True, color=self.pal["cinza"])
            cr.alignment = Alignment(horizontal="left", vertical="top", indent=1)
            ws.row_dimensions[r0].height = 20
            ws.row_dimensions[r0 + 1].height = 16
            ws.row_dimensions[r0 + 2].height = 16
            col = c1 + 2  # uma coluna de respiro entre cartões
        return r0 + 5

    # ---------- notas ----------
    def notas(self, linhas, nome="Notas"):
        """Aba de notas — fonte do dado, premissas, responsável e data.

        Vai por ÚLTIMO e é o que separa uma planilha entregue de uma planilha
        publicada: sem ela, quem recebe não sabe de onde veio o número nem o que
        foi assumido para chegar nele."""
        ws = self.wb.create_sheet(title=_nome_de_aba(nome, "Notas"))
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 2.5
        ws.column_dimensions["B"].width = 110
        c = ws.cell(row=2, column=2, value="Notas, fonte e premissas")
        c.font = Font(name=self.fonte_titulo, size=15, bold=True, color=self.pal["tinta"])
        ws.cell(row=3, column=2).border = Border(
            bottom=Side(style="medium", color=self.pal["tinta"]))
        linha = 5
        for texto in list(linhas):
            cel = ws.cell(row=linha, column=2, value=limpa_texto(texto))
            cel.font = Font(name=self.fonte, size=10.5, color=self.pal["corpo"])
            cel.alignment = Alignment(vertical="top", wrap_text=True)
            linha += 1
        rodape = " · ".join(x for x in (self.emissor, self.cliente,
                                        "Elaborado em " + fmt.data()) if x)
        cel = ws.cell(row=linha + 1, column=2, value=rodape)
        cel.font = Font(name=self.fonte, size=8, bold=True, color=self.pal["cinza"])
        self._notas = ws
        return ws

    # ---------- gráficos ----------
    def grafico_barras(self, ws, info, categoria, valor, titulo="", anchor=None):
        return self._grafico(BarChart(), ws, info, categoria, valor, titulo, anchor)

    def grafico_linhas(self, ws, info, categoria, valor, titulo="", anchor=None):
        return self._grafico(LineChart(), ws, info, categoria, valor, titulo, anchor)

    def grafico_pizza(self, ws, info, categoria, valor, titulo="", anchor=None):
        return self._grafico(PieChart(), ws, info, categoria, valor, titulo, anchor)

    def _grafico_por_tipo(self, tipo, ws, info, categoria, valor, titulo, anchor):
        chart = {"barras": BarChart, "linhas": LineChart,
                 "pizza": PieChart}.get(str(tipo).lower())
        if chart is None:
            raise KitError('painel: tipo de gráfico deve ser barras | linhas | '
                           'pizza (recebi %r)' % (tipo,))
        return self._grafico(chart(), ws, info, categoria, valor, titulo, anchor)

    def _tema(self, chart, npontos, uma_serie):
        """Pinta o gráfico com as cores da identidade.

        `DataPoint(graphicalProperties=...)` levanta TypeError — o argumento do
        construtor é `spPr`; `graphicalProperties` só existe como alias de
        LEITURA. Com o erro engolido em silêncio, a pizza saía com a paleta
        padrão do Excel e ninguém percebia."""
        try:
            if isinstance(chart, PieChart):
                pontos = []
                for i in range(max(0, npontos)):
                    gp = GraphicalProperties(solidFill=CORES_GRAF[i % len(CORES_GRAF)])
                    gp.line.solidFill = "FFFFFF"
                    pontos.append(DataPoint(idx=i, spPr=gp))
                if chart.series:
                    chart.series[0].data_points = pontos
            elif isinstance(chart, LineChart):
                for i, s in enumerate(chart.series):
                    s.graphicalProperties = GraphicalProperties()
                    s.graphicalProperties.line.solidFill = CORES_GRAF[i % len(CORES_GRAF)]
                    s.graphicalProperties.line.width = 28575  # ~2,25 pt
                    s.smooth = False
            else:
                for i, s in enumerate(chart.series):
                    s.graphicalProperties = GraphicalProperties(
                        solidFill=CORES_GRAF[i % len(CORES_GRAF)])
                if hasattr(chart, "gapWidth"):
                    chart.gapWidth = 60
            # Legenda com uma série só repete o título do gráfico.
            if uma_serie and not isinstance(chart, PieChart):
                chart.legend = None
        except Exception as e:
            # O estilo nunca derruba a geração do arquivo — mas o silêncio
            # total já escondeu um gráfico saindo com a cor padrão. Avisa.
            warnings.warn("xlspro: tema do gráfico não aplicado (%r)" % (e,))

    def _grafico(self, chart, ws, info, categoria, valor, titulo, anchor):
        """`ws` é a aba de DESTINO (pode ser o painel); os dados vêm sempre de
        `info["ws"]`, a aba onde a tabela foi escrita."""
        cat_c = self._idx(info, categoria)
        val_c = self._idx(info, valor)
        # Exclui a linha de TOTAL do gráfico: senão ela vira uma fatia/barra
        # gigante (= soma de todas as outras) que distorce a leitura.
        last = info["r1"] - (1 if info.get("total") else 0)
        if last <= info["r0"]:
            raise KitError("grafico_*: a tabela não tem linha de dados para o gráfico")
        dados = Reference(info["ws"], min_col=val_c, min_row=info["r0"], max_row=last)
        cats = Reference(info["ws"], min_col=cat_c, min_row=info["r0"] + 1, max_row=last)
        chart.add_data(dados, titles_from_data=True)
        chart.set_categories(cats)
        chart.title = titulo or valor
        chart.height = 8
        chart.width = 16
        # Eixo de valor SEM centavos: "R$ 600.000,00" em cada marca é ilegível.
        if not isinstance(chart, PieChart):
            try:
                chart.y_axis.numFmt = '#,##0'
                chart.y_axis.majorGridlines.spPr = GraphicalProperties()
            except Exception:                                  # pragma: no cover
                pass
        self._tema(chart, last - info["r0"], uma_serie=True)
        if not anchor:
            anchor = get_column_letter(info["c1"] + 2) + str(info["r0"])
        ws.add_chart(chart, anchor)
        return chart

    # ---------- impressão ----------
    def imprimir(self, ws, orientacao="auto"):
        """Ajusta a aba para caber na LARGURA da folha impressa.

        Sem isto, uma planilha de 12 colunas imprime metade numa folha e o
        resto noutra, com o cabeçalho só na primeira — que foi o defeito do
        painel na revisão da v1."""
        colunas = ws.max_column or 1
        if orientacao == "auto":
            orientacao = "paisagem" if colunas > COLUNAS_PARA_PAISAGEM else "retrato"
        ws.page_setup.orientation = ("landscape" if orientacao == "paisagem"
                                     else "portrait")
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        # `fitToPage` mora em sheetPr/pageSetUpPr: sem ele o Excel IGNORA o
        # fitToWidth acima e imprime em escala 100% assim mesmo.
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins.left = ws.page_margins.right = 0.59   # 1,5 cm
        ws.page_margins.top = ws.page_margins.bottom = 0.59
        ws.oddFooter.right.text = "Página &P de &N"
        ws.oddFooter.right.size = 8
        ws.oddFooter.left.text = limpa_texto(self.emissor or self.titulo_doc or "")
        ws.oddFooter.left.size = 8
        # A área de impressão precisa cobrir também os GRÁFICOS: eles flutuam
        # sobre a grade e não entram em `max_row`/`max_column`, então uma área
        # calculada só pelas células imprime o painel SEM os gráficos — que é
        # justamente o que o painel tem para mostrar.
        linhas, colunas = self._extensao(ws)
        if linhas and colunas:
            ws.print_area = "A1:%s%d" % (get_column_letter(colunas), linhas)
        return ws

    def _extensao(self, ws):
        """`(última linha, última coluna)` da aba, contando os gráficos."""
        linhas, colunas = ws.max_row or 0, ws.max_column or 0
        for chart in getattr(ws, "_charts", []):
            ancora = getattr(chart, "anchor", None)
            if isinstance(ancora, str):
                # `add_chart` guarda a âncora como TEXTO ("B10") e só a converte
                # em objeto na hora de gravar — ler `._from` aqui devolveria
                # None e a área de impressão sairia sem os gráficos.
                from openpyxl.utils.cell import coordinate_to_tuple
                linha0, coluna0 = coordinate_to_tuple(ancora)
            else:
                de = getattr(ancora, "_from", None)
                if de is None:
                    continue
                linha0, coluna0 = de.row + 1, de.col + 1
            # Altura/largura do gráfico vêm em centímetros; uma linha tem ~0,5 cm
            # e uma coluna padrão ~1,8 cm.
            linhas = max(linhas, linha0 + int((chart.height or 8) / 0.5))
            colunas = max(colunas, coluna0 + int((chart.width or 16) / 1.8))
        return linhas, colunas

    # ---------- metadados e saída ----------
    def _metadados(self):
        props = self.wb.properties
        props.title = limpa_texto(self.titulo_doc or "Planilha")
        props.creator = limpa_texto(self.emissor or "Frederico AI Studio")
        props.lastModifiedBy = props.creator
        props.subject = limpa_texto(self.cliente)
        props.language = "pt-BR"
        props.created = props.modified = datetime.now()

    def salvar(self, caminho, auditar=True):
        """Grava, aplica a configuração de impressão em todas as abas e AUDITA.

        Devolve `{"ok", "abas", "achados"}`. Achado grave levanta `KitError`."""
        if not self.wb.worksheets:
            raise KitError("xlspro: a pasta de trabalho não tem nenhuma aba")
        if self._notas is not None:
            # A aba de notas fecha o arquivo: move para a última posição caso
            # outra aba tenha nascido depois dela.
            self.wb.move_sheet(self._notas, offset=len(self.wb.worksheets))
        for ws in self.wb.worksheets:
            self.imprimir(ws)
        self._metadados()
        self.wb.save(caminho)
        if not auditar:
            return relatorio([], abas=len(self.wb.worksheets))
        rel = relatorio(self._auditar(caminho), abas=len(self.wb.worksheets))
        return falha_se_grave(rel, "xlspro")

    # ---------- auditoria ----------
    def _auditar(self, caminho):
        from openpyxl import load_workbook
        achados = list(self._achados_geracao)
        wb = load_workbook(caminho)
        textos = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                textos += [v for v in row if isinstance(v, str)]
        achados += achados_de_placeholder(textos, "planilha")
        achados += self._auditar_tabelas(wb)
        if self._painel is not None and wb.worksheets[0].title != self._painel.title:
            achados.append(achado("grave", "painel-fora-do-lugar",
                                  "o painel deveria ser a PRIMEIRA aba, mas a "
                                  "primeira é %r" % wb.worksheets[0].title))
        if self._notas is not None and wb.worksheets[-1].title != self._notas.title:
            achados.append(achado("aviso", "notas-fora-do-lugar",
                                  "a aba de notas deveria ser a última"))
        for ws in wb.worksheets:
            if ws.sheet_properties.pageSetUpPr is None or \
                    not ws.sheet_properties.pageSetUpPr.fitToPage:
                achados.append(achado("aviso", "impressao-sem-ajuste",
                                      "a aba %r não está ajustada à largura da "
                                      "folha impressa" % ws.title))
        achados += self._auditar_graficos(caminho, wb)
        achados += self._auditar_formulas(caminho, wb)
        return achados

    def _auditar_tabelas(self, wb):
        """Coluna declarada numérica que recebeu TEXTO — é o que faz a soma dar
        zero e o gráfico sair vazio, sem nenhum erro visível na tela."""
        achados = []
        for info in self._tabelas:
            try:
                ws = wb[info["ws"].title]
            except KeyError:                                   # pragma: no cover
                continue
            ultima = info["r1"] - (1 if info["total"] else 0)
            for j, tipo in info["tipos"].items():
                if tipo == "data":
                    continue
                for r in range(info["r0"] + 1, ultima + 1):
                    v = ws.cell(row=r, column=j + 1).value
                    if v in (None, ""):
                        continue
                    if isinstance(v, str) and not v.startswith("="):
                        achados.append(achado(
                            "grave", "coluna-numerica-com-texto",
                            "aba %r, célula %s%d: a coluna %r é %s e recebeu o "
                            "texto %r — passe o NÚMERO, o kit é quem formata"
                            % (ws.title, get_column_letter(j + 1), r,
                               info["cols"][j], tipo, v[:24])))
                        break
        return achados

    def _auditar_graficos(self, caminho, wb):
        """Reaproveita a validação de gráfico do `validar_artefato` (aba
        inexistente, intervalo invertido, série vazia) em vez de reescrevê-la."""
        try:
            import validar_artefato
        except Exception:                                      # pragma: no cover
            return []
        try:
            _n, problemas = validar_artefato.check_charts(caminho, wb)
        except Exception as e:                                 # pragma: no cover
            return [achado("aviso", "grafico-nao-conferido",
                           "não foi possível conferir os gráficos (%r)" % (e,))]
        return [achado("grave", "grafico-invalido", p) for p in problemas]

    def _auditar_formulas(self, caminho, wb):
        """Só quando há fórmula: recalcula com o LibreOffice e procura #REF!,
        #DIV/0! e companhia. Sem recálculo, o erro só aparece no Excel do
        cliente — que é tarde demais."""
        tem_formula = any(isinstance(c.value, str) and c.value.startswith("=")
                          for ws in wb.worksheets for row in ws.iter_rows()
                          for c in row)
        if not tem_formula:
            return []
        try:
            import shutil
            import validar_artefato
            cfg = validar_artefato.Config.from_env()
            calc, tmp = validar_artefato.recalc(caminho, cfg)
            if not calc:
                return [achado("aviso", "formula-nao-recalculada",
                               "há fórmulas e não foi possível recalcular "
                               "(LibreOffice indisponível)")]
            from openpyxl import load_workbook
            erros, _capped = validar_artefato.scan_errors(
                load_workbook(calc, data_only=True), cfg.max_cells)
            shutil.rmtree(tmp, ignore_errors=True)
        except Exception as e:                                 # pragma: no cover
            return [achado("aviso", "formula-nao-recalculada",
                           "não foi possível recalcular as fórmulas (%r)" % (e,))]
        if erros:
            return [achado("grave", "formula-com-erro",
                           "%d célula(s) com erro de fórmula depois do "
                           "recálculo" % erros)]
        return []
